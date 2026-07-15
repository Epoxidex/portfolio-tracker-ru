from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_
from .. import config
from ..db import get_db
from ..models import Instrument, Transaction
from ..schemas import (
    TxIn, InstrumentIn, PriceIn, DepositIn, CurrencyHoldingIn, TrackingStartIn,
    BackupRestoreIn, CashLedgerIn, DepositOpenLedgerIn, DepositSettleLedgerIn,
    CurrencyBuyLedgerIn, CurrencySellLedgerIn, SecurityBuyLedgerIn,
    SecuritySellLedgerIn, LedgerBatchIn,
)
from ..dataio import DATABASE_MAINTENANCE_LOCK, backup_database
from ..services import portfolio, calendar as cal, snapshots, read_model, ledger
from ..services.onboarding import SetupConflict, add_currency_holding, create_deposit
from ..services.tracking import apply_tracking_cleanup, update_env_setting
from ..services.tinvest import fetch_prices
from ..services.banki import fetch_fx
from ..services.operations import sync_operations
from ..services.git_backup import (
    GitBackupError,
    create_repository_backup,
    list_repository_backups,
    restore_repository_backup,
)

router = APIRouter(prefix="/api")


def _resolve(db, instrument_id, ticker):
    if instrument_id:
        return db.get(Instrument, instrument_id)
    if ticker:
        return db.query(Instrument).filter(
            or_(Instrument.ticker == ticker, Instrument.isin == ticker)).first()
    return None


@router.get("/summary")
def get_summary(db: Session = Depends(get_db)):
    return portfolio.summary(db)


@router.get("/status")
def get_status(db: Session = Depends(get_db)):
    """Configuration and data readiness without exposing secrets or local paths."""
    return read_model.data_status(db)


@router.get("/backups")
def get_backups():
    try:
        return {"ok": True, "items": list_repository_backups()}
    except GitBackupError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@router.post("/backups")
def create_backup():
    try:
        return {"ok": True, "backup": create_repository_backup()}
    except GitBackupError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@router.post("/backups/restore")
def restore_backup(payload: BackupRestoreIn):
    try:
        result = restore_repository_backup(payload.filename)
        return {"ok": True, **result}
    except GitBackupError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@router.post("/settings/tracking-start")
def set_tracking_start(payload: TrackingStartIn, db: Session = Depends(get_db)):
    """Back up the DB, trim older broker imports, and persist the tracking date."""
    backup = backup_database(prefix="before-tracking-start")
    result = apply_tracking_cleanup(db, payload.start_date)
    update_env_setting(
        config.BASE_DIR / ".env",
        "PORTFOLIO_TRACKING_START_DATE",
        payload.start_date.isoformat(),
    )
    config.PORTFOLIO_TRACKING_START_DATE = payload.start_date
    snapshot = snapshots.take_snapshot(db, source="tracking-reset")
    return {
        "ok": True,
        "start_date": payload.start_date.isoformat(),
        "deleted": {
            "transactions": result["imported_transactions"],
            "instruments": result["instruments"],
            "snapshots": result["snapshots"],
        },
        "backup": backup.name,
        "snapshot": snapshot.ts.isoformat(),
    }


@router.get("/positions")
def get_positions(db: Session = Depends(get_db)):
    return portfolio.positions(db)


@router.get("/calendar")
def get_calendar(months: int = 24, past: bool = False, db: Session = Depends(get_db)):
    return cal.calendar(db, months_ahead=months, include_past=past)


@router.get("/income")
def get_income(db: Session = Depends(get_db)):
    return cal.passive_income(db)


@router.get("/ledger/cash")
def get_ledger_cash(db: Session = Depends(get_db)):
    return ledger.rub_cash_balance(db)


@router.get("/ledger/realized")
def get_ledger_realized(db: Session = Depends(get_db)):
    return portfolio.realized_results(db)


@router.get("/ledger/reconciliations")
def get_ledger_reconciliations(db: Session = Depends(get_db)):
    return ledger.pending_reconciliations(db)


def _ledger_payload(payload, action_type: str) -> tuple[str, bool, dict]:
    data = payload.model_dump()
    request_id = data.pop("request_id")
    data.pop("confirm", None)
    create_snapshot = data.pop("create_snapshot", True)
    data["type"] = action_type
    return request_id, create_snapshot, data


def _apply_ledger(db: Session, request_id: str, actions: list[dict], create_snapshot: bool):
    try:
        with DATABASE_MAINTENANCE_LOCK:
            return ledger.apply_actions(
                db,
                request_id=request_id,
                actions=actions,
                create_snapshot=create_snapshot,
            )
    except ledger.LedgerConflict as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@router.post("/ledger/cash/topup")
def ledger_cash_topup(payload: CashLedgerIn, db: Session = Depends(get_db)):
    request_id, create_snapshot, action = _ledger_payload(payload, "cash_topup")
    return _apply_ledger(db, request_id, [action], create_snapshot)


@router.post("/ledger/cash/withdrawal")
def ledger_cash_withdrawal(payload: CashLedgerIn, db: Session = Depends(get_db)):
    request_id, create_snapshot, action = _ledger_payload(payload, "cash_withdrawal")
    return _apply_ledger(db, request_id, [action], create_snapshot)


@router.post("/ledger/deposits/open")
def ledger_open_deposit(payload: DepositOpenLedgerIn, db: Session = Depends(get_db)):
    request_id, create_snapshot, action = _ledger_payload(payload, "open_deposit")
    return _apply_ledger(db, request_id, [action], create_snapshot)


@router.post("/ledger/deposits/settle")
def ledger_settle_deposit(payload: DepositSettleLedgerIn, db: Session = Depends(get_db)):
    request_id, create_snapshot, action = _ledger_payload(payload, "settle_deposit")
    return _apply_ledger(db, request_id, [action], create_snapshot)


@router.post("/ledger/currencies/buy")
def ledger_buy_currency(payload: CurrencyBuyLedgerIn, db: Session = Depends(get_db)):
    request_id, create_snapshot, action = _ledger_payload(payload, "buy_currency")
    return _apply_ledger(db, request_id, [action], create_snapshot)


@router.post("/ledger/currencies/sell")
def ledger_sell_currency(payload: CurrencySellLedgerIn, db: Session = Depends(get_db)):
    request_id, create_snapshot, action = _ledger_payload(payload, "sell_currency")
    return _apply_ledger(db, request_id, [action], create_snapshot)


@router.post("/ledger/securities/buy")
def ledger_buy_security(payload: SecurityBuyLedgerIn, db: Session = Depends(get_db)):
    request_id, create_snapshot, action = _ledger_payload(payload, "buy_security")
    return _apply_ledger(db, request_id, [action], create_snapshot)


@router.post("/ledger/securities/sell")
def ledger_sell_security(payload: SecuritySellLedgerIn, db: Session = Depends(get_db)):
    request_id, create_snapshot, action = _ledger_payload(payload, "sell_security")
    return _apply_ledger(db, request_id, [action], create_snapshot)


@router.post("/ledger/actions")
def ledger_actions(payload: LedgerBatchIn, db: Session = Depends(get_db)):
    return _apply_ledger(
        db,
        payload.request_id,
        payload.model_dump()["actions"],
        payload.create_snapshot,
    )


@router.get("/history")
def get_history(db: Session = Depends(get_db)):
    return snapshots.history(db)


@router.get("/instruments")
def list_instruments(db: Session = Depends(get_db)):
    return read_model.list_instruments(db, limit=2000)["items"]


@router.post("/instruments")
def add_instrument(payload: InstrumentIn, db: Session = Depends(get_db)):
    inst = Instrument(**payload.model_dump())
    db.add(inst); db.commit(); db.refresh(inst)
    return {"id": inst.id}


@router.post("/deposits")
def add_deposit(payload: DepositIn, db: Session = Depends(get_db)):
    """Create the deposit and its opening cash flow in one transaction."""
    try:
        return create_deposit(db, **payload.model_dump())
    except SetupConflict as exc:
        raise HTTPException(409, str(exc)) from exc


@router.post("/currencies")
def add_currency(payload: CurrencyHoldingIn, db: Session = Depends(get_db)):
    """Add an opening manual currency balance without requiring raw SQL/API objects."""
    try:
        return add_currency_holding(db, **payload.model_dump())
    except SetupConflict as exc:
        raise HTTPException(409, str(exc)) from exc


@router.post("/transactions")
def add_transaction(payload: TxIn, db: Session = Depends(get_db)):
    data = payload.model_dump()
    ticker = data.pop("ticker", None)
    inst = _resolve(db, data.get("instrument_id"), ticker)
    requires_instrument = {"buy", "sell", "coupon", "dividend", "fx_buy", "fx_sell"}
    if data["kind"] in requires_instrument and inst is None:
        raise HTTPException(400, "instrument not found")
    data["instrument_id"] = inst.id if inst else data.get("instrument_id")
    tx = Transaction(**data)
    db.add(tx); db.commit(); db.refresh(tx)
    return {"id": tx.id}


@router.delete("/transactions/{tx_id}")
def del_transaction(tx_id: int, db: Session = Depends(get_db)):
    tx = db.get(Transaction, tx_id)
    if not tx:
        raise HTTPException(404)
    db.delete(tx); db.commit()
    return {"ok": True}


@router.get("/transactions")
def list_transactions(db: Session = Depends(get_db)):
    return read_model.list_transactions(db, descending=False, limit=2000)["items"]


@router.post("/price")
def set_price(payload: PriceIn, db: Session = Depends(get_db)):
    inst = _resolve(db, payload.instrument_id, payload.ticker)
    if not inst:
        raise HTTPException(404, "instrument not found")
    inst.last_price = payload.last_price
    if payload.nkd is not None:
        inst.nkd = payload.nkd
    db.commit()
    return {"ok": True, "instrument": inst.name, "last_price": inst.last_price}


@router.post("/snapshot")
def make_snapshot(db: Session = Depends(get_db)):
    s = snapshots.take_snapshot(db, source="manual")
    return {"ok": True, "ts": s.ts.isoformat(), "value": s.total_value}


@router.post("/fetch/prices")
def do_fetch_prices(db: Session = Depends(get_db)):
    return fetch_prices(db)


@router.post("/fetch/fx")
def do_fetch_fx(
    source: str | None = Query(default=None, description="bank_buy | bank_sell | cbr"),
    db: Session = Depends(get_db),
):
    """Обновляет курсы валют. source переопределяет FX_RATE_SOURCE из .env."""
    return fetch_fx(db, source=source)


@router.get("/prices/history")
def get_price_history(days: int = Query(default=90), db: Session = Depends(get_db)):
    """История цен всех инструментов за последние N дней."""
    return read_model.price_history(db, days=days, limit_per_instrument=10000)["items"]


@router.post("/sync/operations")
def do_sync_operations(
    days: int = Query(default=365, description="глубина истории в днях"),
    db: Session = Depends(get_db),
):
    """Импортирует операции из T-Invest (buy/sell/coupon/dividend)."""
    return sync_operations(db, days_back=days)


@router.post("/sync/tinvest")
def do_sync_tinvest(
    days: int = Query(default=3650, ge=1, le=36500, description="глубина истории в днях"),
    db: Session = Depends(get_db),
):
    """Import supported operations, refresh prices, then save a snapshot."""
    if not config.TINVEST_TOKEN:
        raise HTTPException(400, "TINVEST_TOKEN не задан в .env")
    operations = sync_operations(db, days_back=days)
    if not operations.get("ok"):
        raise HTTPException(502, operations.get("error", "не удалось импортировать операции"))
    prices = fetch_prices(db)
    if not prices.get("ok"):
        raise HTTPException(502, prices.get("error", "не удалось обновить цены"))
    snap = snapshots.take_snapshot(db, source="tinvest-sync")
    return {
        "ok": True,
        "imported": operations.get("imported", 0),
        "skipped": operations.get("skipped", 0),
        "prices_updated": len(prices.get("updated", [])),
        "warnings": prices.get("warnings", []),
        "snapshot": snap.ts.isoformat(),
    }


# ---------- Доходность по периодам ----------

@router.get("/returns")
def get_returns(period: str = Query(default="monthly"), db: Session = Depends(get_db)):
    return snapshots.compute_returns(db, period=period)


@router.get("/leaders")
def get_leaders(
    period: str = Query(default="day", pattern="^(day|week|month)$"),
    db: Session = Depends(get_db),
):
    return snapshots.compute_leaders(db, period=period)


# ---------- Экспорт в Excel ----------

@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db)):
    import openpyxl, io
    from datetime import date as dt_date

    wb = openpyxl.Workbook()

    # Лист 1: Позиции
    ws = wb.active
    ws.title = "Позиции"
    pos_headers = ["Название", "Тип", "Кол-во", "Вложено ₽", "Стоимость ₽",
                   "Доход ₽", "P&L ₽", "P&L %", "Цена", "НКД"]
    ws.append(pos_headers)
    for p in portfolio.positions(db):
        ws.append([
            p["name"], p["kind"], p["qty"], p["invested"], p["value"],
            p["income"], p["pnl"], round(p["pnl_pct"] * 100, 2),
            p["last_price"], p["nkd"],
        ])

    # Лист 2: Транзакции
    ws2 = wb.create_sheet("Транзакции")
    tx_headers = ["Дата", "Инструмент", "Тип", "Кол-во", "Сумма ₽", "Комиссия", "Заметка"]
    ws2.append(tx_headers)
    for t in db.query(Transaction).order_by(Transaction.ts).all():
        ws2.append([
            t.ts.isoformat(),
            t.instrument.name if t.instrument else "",
            t.kind, t.quantity, t.amount, t.commission, t.note,
        ])

    # Лист 3: Снапшоты
    ws3 = wb.create_sheet("Снапшоты")
    ws3.append(["Дата", "Стоимость ₽", "Вложено ₽", "P&L ₽", "Доход ₽"])
    for row in snapshots.history(db):
        ws3.append([row["ts"][:10], row["value"], row["invested"],
                    row["pnl"], row["income"]])

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    fname = f"portfolio_{dt_date.today().isoformat()}.xlsx"
    from fastapi.responses import Response as _Resp
    return _Resp(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "Content-Length": str(len(content)),
        },
    )
